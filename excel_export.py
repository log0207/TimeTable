from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
import io

def export_timetable_to_excel(timetable_manager, shift, filename, department_id=None):
    wb = Workbook()
    ws = wb.active
    
    department_name = "All Departments"
    if department_id:
        department = timetable_manager.db.get_department(department_id)
        department_name = department[1] if department else "Unknown Department"
    
    ws.title = f"{timetable_manager.shifts[shift]['name']} - {department_name}"

    # Add header with shift and department information
    ws.merge_cells('A1:I1')
    header = ws.cell(row=1, column=1, value=f"Timetable - {timetable_manager.shifts[shift]['name']} - {department_name}")
    header.font = Font(bold=True, size=14)
    header.alignment = Alignment(horizontal='center')

    timetable_data = timetable_manager.get_timetable_data(shift, department_id)

    if department_id:
        export_single_department(ws, timetable_data, timetable_manager.shifts[shift])
    else:
        export_all_departments(wb, timetable_data, timetable_manager.shifts[shift])

    wb.save(filename)
    print(f"Timetable exported to {filename}")

def export_single_department(ws, timetable_data, shift_info):
    periods = shift_info['periods']

    # Add period numbers
    for i in range(1, len(periods) + 1):
        ws.cell(row=2, column=i+1, value=i)
        ws.cell(row=2, column=i+1).font = Font(bold=True, size=14)
        ws.cell(row=2, column=i+1).alignment = Alignment(horizontal='center')

    # Add period details
    for i, period in enumerate(periods, start=1):
        ws.cell(row=3, column=i+1, value=f"{period['type'].capitalize()}\n({period['start']}-{period['end']})")
        ws.cell(row=3, column=i+1).font = Font(bold=True)
        ws.cell(row=3, column=i+1).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    row = 4

    for entry in timetable_data:
        day = entry['Day']
        ws.cell(row=row, column=1, value=day)

        for period in entry.get('Periods', []):
            period_num = period['Period']
            if period['Type'] == 'regular':
                course = period['Course'] if period['Course'] else ''
                staff = period['Staff'] if period['Staff'] else ''
                cell_value = f"{course}\n{staff}" if course or staff else ''
            else:
                cell_value = period['Type'].capitalize()
            
            cell = ws.cell(row=row, column=period_num+1, value=cell_value)
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

        row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    for col in range(2, len(periods) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Add borders
    add_borders(ws, row-1, len(periods)+1)

def export_all_departments(wb, timetable_data, shift_info):
    for dept_data in timetable_data:
        ws = wb.create_sheet(title=dept_data['department_name'])
        export_single_department(ws, dept_data['timetable'], shift_info)

    # Remove the default sheet created
    wb.remove(wb['Sheet'])

def add_borders(ws, max_row, max_col):
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    for row in ws[f'A1:{get_column_letter(max_col)}{max_row}']:
        for cell in row:
            cell.border = thin_border

def export_timetable_as_image(timetable_manager, shift, filename, department_id=None):
    timetable_data = timetable_manager.get_timetable_data(shift, department_id)
    
    fig, ax = plt.subplots(figsize=(20, 10))
    ax.axis('tight')
    ax.axis('off')
    
    table_data = []
    periods = timetable_manager.shifts[shift]['periods']
    
    # Add period numbers
    period_numbers = ['Day'] + [str(i) for i in range(1, len(periods) + 1)]
    table_data.append(period_numbers)
    
    # Add period details
    period_details = ['']
    for i, period in enumerate(periods, start=1):
        if period['type'] == 'regular':
            period_details.append(f"Period {i}\n({period['start']}-{period['end']})")
        else:
            period_details.append(f"{period['type'].capitalize()}\n({period['start']}-{period['end']})")
    table_data.append(period_details)
    
    for day_data in timetable_data:
        row = [day_data['Day']]
        for period in day_data['Periods']:
            if period['Type'] == 'regular':
                cell_text = f"{period['Course']}\n{period['Staff']}"
            else:
                cell_text = period['Type'].capitalize()
            row.append(cell_text)
        table_data.append(row)
    
    table = ax.table(cellText=table_data, cellLoc='center', loc='center')
    table.auto_set_font_size(False)
    table.set_fontsize(8)
    table.scale(1, 2)
    
    # Customize cell colors
    for i in range(len(table_data)):
        for j in range(len(table_data[0])):
            cell = table[i, j]
            if i < 2:  # Header rows
                cell.set_facecolor('#ADD8E6')  # Light blue
            elif j == 0:  # Day column
                cell.set_facecolor('#F0F0F0')  # Light gray
    
    department_name = timetable_manager.db.get_department(department_id)[1] if department_id else "All Departments"
    plt.title(f"Timetable - {timetable_manager.shifts[shift]['name']} - {department_name}")
    plt.savefig(filename, dpi=300, bbox_inches='tight')
    print(f"Timetable image exported to {filename}")

def export_timetable(timetable_manager, shift, filename, department_id=None):
    export_timetable_to_excel(timetable_manager, shift, filename + '.xlsx', department_id)
    export_timetable_as_image(timetable_manager, shift, filename + '.png', department_id)
    print(f"Timetable exported to {filename}.xlsx and {filename}.png")